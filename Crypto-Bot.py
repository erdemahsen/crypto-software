import openpyxl
from openpyxl import *
from datetime import *


class Person:
    def __init__(self, startingDate, endingDate, subPeriod,email, tw, tg, dc, txid, coinType, amount, expired) :
        self.startingDate = startingDate
        self.endingDate = endingDate
        self.subPeriod = subPeriod
        self.email = email
        self.tw = tw
        self.tg = tg
        self.dc = dc
        self.txid = txid
        self.coinType = coinType
        self.amount = amount
        self.expire = expired

def dateConvert(DateString): #Gets a date, example:2021-11-23 18:58:51
    date = DateString.split(" ")
    date = date[0].split("-")
    year = int(date[0])
    month = int(date[1])
    day = int(date[2])
    date = datetime.today()
    date = date.replace(year = year, month = month, day = day)
    #print(date)
    return date

def dateConvertToString(date):
    year = str(date.year)
    month = str(date.month)
    day = str(date.day)
    newstr = day + "/" + month + "/" + year
    return newstr

#Test Place
dateConvert("2021-11-23 18:58:51")
#Test Place
def dateForward(date, dayForward): #Gets a tuple in the form of (day, month, year)
    intyear = int(date.year)
    intmonth = int(date.month)
    intday = int(date.day)
    for i in range(dayForward):
        maxdays = 0
        if(intmonth == 2):
            maxdays = 28
            if(intyear / 4 == int(intyear/4)):
                maxdays = 29
        elif (intmonth == 1 or intmonth == 3 or intmonth == 5 or intmonth == 7 or intmonth == 8 or intmonth == 10 or intmonth == 12):
            maxdays = 31
        else:
            maxdays = 30
        intday = intday+1
        if (intday > maxdays) :
            intday = 1
            intmonth = intmonth + 1
        if (intmonth > 12) :
            intmonth = 1
            intyear = intyear + 1
    date = datetime.today()
    date = date.replace(year = intyear, month = intmonth, day = intday)
    #print(date)
    return date

#Test Place
dateForward(dateConvert("2021-11-23 18:58:51"), 1923)    
#Test Place
personList = []
sizePeople = 1
depositSize = 0
googleSize = 0
deposit = load_workbook(filename = "deposit_history.xlsx")
HistorySheet = deposit.worksheets[0]
GoogleSheet = deposit.worksheets[1]

History_Date_A = HistorySheet["A"]
History_Coin_B = HistorySheet["B"]
History_Amount_C = HistorySheet["C"]
History_Address_E = HistorySheet["E"]
History_Txid_F = HistorySheet["F"]
History_Status_I = HistorySheet["I"]

Google_Date_A = GoogleSheet["A"]
Google_Email_B = GoogleSheet["B"]
Google_Tw_C = GoogleSheet["C"]
Google_Tg_D = GoogleSheet["D"]
Google_Dc_E = GoogleSheet["E"]
Google_Txid_F = GoogleSheet["F"]

for i in History_Txid_F:
    depositSize += 1
for i in Google_Txid_F:
    googleSize += 1

workbook = Workbook()
sheet = workbook.active
sheet["A1"] = "Starting Date"
sheet["B1"] = "Ending Date"
sheet["C1"] = "Sub Period"
sheet["D1"] = "Twitter"
sheet["E1"] = "Telegram"
sheet["F1"] = "Discord"
sheet["G1"] = "TXID"
sheet["H1"] = "Coin Type"
sheet["I1"] = "Amount"
sheet["J1"] = "Expire Condition"
sheet["K1"] = "Email"

txidused = []
alltxid = []
txidnotused = []

for i in range(1, depositSize):

    History_Date_A_Cell = History_Date_A[i]
    History_Coin_B_Cell = History_Coin_B[i]
    History_Amount_C_Cell = History_Amount_C[i]
    History_Status_I_Cell = History_Status_I[i]   
    History_Address_E_Cell = History_Address_E[i]
    History_Txid_F_Cell = History_Txid_F[i]
    alltxid.append(History_Txid_F_Cell.value)
    for j in range(1, googleSize):
        
        txidflag = 0
        Google_Txid_F_Cell = Google_Txid_F[j]
        flag = 1
        googletxid = Google_Txid_F_Cell.value
        
        historytxidlist = History_Txid_F_Cell.value.split(" ")
        historytxid = historytxidlist[len(historytxidlist)-1]
        if((Google_Txid_F_Cell.value == History_Txid_F_Cell.value or historytxid == str(googletxid) )and History_Status_I_Cell.value == "Completed"):
            
            Google_Date_A_Cell = Google_Date_A[j]
            Google_Email_B_Cell = Google_Email_B[j]
            Google_Tw_C_Cell = Google_Tw_C[j]
            Google_Tg_D_Cell = Google_Tg_D[j]
            Google_Dc_E_Cell = Google_Dc_E[j]        

            flag = 0
            for tx in txidused:
                if (tx == Google_Txid_F_Cell.value):
                    flag = 1
            
            timestamp = 0
            paymentRequired = -1
            TRC20USDT = 28
            BEP20USDT = 28
            BEP20BUSD = 28
            BEP20address = "adressfield" #this is the place you should edit
            amount = History_Amount_C_Cell.value
            coinType = History_Coin_B_Cell.value

            if(coinType == "USDT"):
                if(History_Address_E_Cell.value == BEP20address):
                    paymentRequired = BEP20USDT
                else :
                    paymentRequired = TRC20USDT
            elif(coinType == "BUSD"):
                paymentRequired = BEP20BUSD
            timestamp = (float(amount) / paymentRequired)
            timestamp = int(timestamp)
            if (timestamp <= 0):
                flag = 1

            person = Person
            
            startingDate = dateConvert(History_Date_A_Cell.value)
            endingDate = dateForward(dateConvert(History_Date_A_Cell.value), timestamp * 30)
            now =datetime.now()

            person.startingDate = dateConvertToString(startingDate)
            person.endingDate = dateConvertToString(endingDate)
            person.subPeriod = str(timestamp) + "ay"
            person.email = Google_Email_B_Cell.value
            person.tw = Google_Tw_C_Cell.value
            person.tg = Google_Tg_D_Cell.value
            person.dc = Google_Dc_E_Cell.value
            person.txid = History_Txid_F_Cell.value
            person.coinType = History_Coin_B_Cell.value
            person.amount = History_Amount_C_Cell.value
            if(endingDate < now):
                person.expire = "Expired"
            else:
                person.expire = "Continues"
            
            if (not flag):
                personList.append(person)
                sizePeople+=1
                stri = str(sizePeople)
                txidused.append(person.txid)
                sheet["A"+stri] = person.startingDate
                sheet["B"+stri] = person.endingDate
                sheet["C"+stri] = person.subPeriod
                sheet["D"+stri] = person.tw
                sheet["E"+stri] = person.tg
                sheet["F"+stri] = person.dc
                sheet["G"+stri] = person.txid
                sheet["H"+stri] = person.coinType
                sheet["I"+stri] = person.amount
                sheet["J"+stri] = person.expire
                sheet["K"+stri] = person.email


txidnotused = list(set(alltxid)-set(txidused))
notusedsheet = Workbook()
sheet = notusedsheet.active
sheet["A1"] = "Not used"
for i in range(1, len(txidnotused)):
    sheet["A"+str(i+1)] = txidnotused[i-1]
notusedsheet.save(filename = "notused.xlsx")
workbook.save(filename="members.xlsx")